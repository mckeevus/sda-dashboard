"""
SDA Dashboard Data Loader v4
==============================
Extends v3 with participant and financial data tables:

  Table P.9  — Participant need at SA4 level (all quarters, era-aware)
  Table P.10 — Participant need by Design Category at SA4 (2024-25 Q2+)
  Table P.17 — Participant need at SA3 level (all quarters, era-aware)
  Table P.18 — Participant need by Design Category at SA3 (2024-25 Q2+)
  Table P.2  — Annualised committed supports / financial (all quarters)

ERA RULES
─────────────────────────────────────────────────────────────────
P.9 / P.17 total column:
  All quarters  →  load canonical 'Total participants' measure
  2024-25 Q1+   →  also load 'Participants with SDA in use' and
                    'Participants SDA eligible, not yet using SDA'

P.10 / P.18:
  Only loaded from 2024-25 Q2 onward (9 quarters of clean DC data).
  'Basic' column exists only in 2024-25 Q2 — loaded where present.

P.2 financial:
  All 13 quarters. Early quarters store dollar values as '$333,693,511'
  (string with $ and commas) — stripped on parse.
  Percentage columns skipped (calculated in dashboard JS).
─────────────────────────────────────────────────────────────────

Usage:
    python load_supplement_p_v4.py                      # Full rebuild
    python load_supplement_p_v4.py --folder /path/to    # Custom folder
    python load_supplement_p_v4.py --verify              # Just verify existing DB
"""

import sqlite3
import os
import re
import sys
import argparse
from pathlib import Path
import pandas as pd
import openpyxl

# ============================================================
# TITLE-BASED SHEET DETECTION
# ============================================================

# Each concept is identified by keywords in the sheet title row.
# The tuple is (must_contain_all, must_not_contain_any, canonical_name)
SHEET_SIGNATURES = [
    # Multi-dimensional tables (PRIMARY - the ones we really want)
    {
        'canonical': 'Table P.11',
        'desc': 'New Build by Build Type x Design Category (SA4)',
        'must_have': ['New Build', 'Build Type'],
        'must_not': ['Unfinished', 'SA3'],
    },
    {
        'canonical': 'Table P.12',
        'desc': 'Existing Stock by Build Type x Design Category (SA4)',
        'must_have': ['Existing Stock', 'Build Type'],
        'must_not': ['In-kind Existing', 'SA3'],
    },
    # Aggregate tables
    {
        'canonical': 'Table P.4',
        'desc': 'Enrolled Dwellings by Building Type (SA4)',
        'must_have': ['Enrolled', 'Building Type'],
        'must_not': ['SA3', 'Design'],
    },
    {
        'canonical': 'Table P.5',
        'desc': 'Enrolled Dwellings by Design Category (SA4)',
        'must_have': ['Enrolled', 'Design Category'],
        'must_not': ['SA3', 'Building Type', 'Maximum'],
    },
    {
        'canonical': 'Table P.6',
        'desc': 'Enrolled Dwellings by Max Residents (SA4)',
        'must_have': ['Enrolled', 'Maximum'],
        'must_not': ['SA3'],
    },
    {
        'canonical': 'Table P.7',
        'desc': 'New Build Max Residents by Design Category (SA4)',
        'must_have': ['New Build', 'Maximum Resident', 'Design Category'],
        'must_not': ['SA3', 'Unfinished'],
    },
    {
        'canonical': 'Table P.8',
        'desc': 'Unfinished New Build by Design Category (SA4)',
        'must_have': ['Unfinished', 'Design Category'],
        'must_not': ['SA3', 'Build Type'],
    },
    # Providers
    {
        'canonical': 'Table P.3',
        'desc': 'Active SDA providers',
        'must_have': ['Active', 'provider'],
        'must_not': [],
    },
    # Participant need SA4 — P.9 canonical (was Table11/P.11/P.9 across eras)
    {
        'canonical': 'Table P.9',
        'desc': 'Participants with identified SDA needs (SA4)',
        'must_have': ['Participants', 'SDA need'],
        'must_not': ['SA3', 'Design Category'],
    },
    # Participant need by Design Category SA4 — P.10 canonical (2024-25 Q2+)
    {
        'canonical': 'Table P.10',
        'desc': 'Participants with SDA need by Design Category (SA4)',
        'must_have': ['Participants', 'SDA need', 'Design Category'],
        'must_not': ['SA3'],
    },
    # Participant need SA3 — P.17 canonical
    {
        'canonical': 'Table P.17',
        'desc': 'Participants with identified SDA needs (SA3)',
        'must_have': ['Participants', 'SDA need', 'SA3'],
        'must_not': ['Design Category'],
    },
    # Participant need by Design Category SA3 — P.18 canonical (2024-25 Q2+)
    {
        'canonical': 'Table P.18',
        'desc': 'Participants with SDA need by Design Category (SA3)',
        'must_have': ['Participants', 'SDA need', 'Design Category', 'SA3'],
        'must_not': [],
    },
    # Financial — P.2 (all quarters)
    {
        'canonical': 'Table P.2',
        'desc': 'Annualised committed supports (SDA and SIL)',
        'must_have': ['Annualised committed support'],
        'must_not': [],
    },
    {
        'canonical': 'Table P.16',
        'desc': 'Unfinished by Build Type x Design Category (SA4)',
        'must_have': ['Unfinished', 'Build Type'],
        'must_not': ['SA3'],
    },
    # SA3 tables
    {
        'canonical': 'Table P.13',
        'desc': 'Enrolled Dwellings by Build Type (SA3)',
        'must_have': ['Enrolled', 'SA3', 'Building Type'],
        'must_not': ['Design'],
    },
    {
        'canonical': 'Table P.14',
        'desc': 'Enrolled Dwellings by Design Category (SA3)',
        'must_have': ['Enrolled', 'SA3', 'Design Category'],
        'must_not': [],
    },
    {
        'canonical': 'Table P.15',
        'desc': 'Enrolled Dwellings by Max Residents (SA3)',
        'must_have': ['Enrolled', 'SA3', 'Maximum'],
        'must_not': [],
    },
    {
        'canonical': 'Table P.17',
        'desc': 'Participants with SDA needs (SA3)',
        'must_have': ['Participants', 'identified SDA need', 'SA3'],
        'must_not': [],
    },
]


def identify_sheet(title):
    """Match a sheet title to a canonical table name using keyword signatures."""
    for sig in SHEET_SIGNATURES:
        if all(kw in title for kw in sig['must_have']):
            if not any(kw in title for kw in sig['must_not']):
                return sig['canonical'], sig['desc']
    return None, None


def detect_sheets(filepath):
    """Scan all sheets in a file, return {canonical_name: sheet_name} mapping."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    mapping = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith('Table'):
            continue
        ws = wb[sheet_name]
        title = ''
        for row in ws.iter_rows(min_row=1, max_row=3, max_col=1, values_only=True):
            if row[0] and len(str(row[0])) > 20:
                title = str(row[0])
                break

        if title:
            canonical, desc = identify_sheet(title)
            if canonical and canonical not in mapping:
                mapping[canonical] = sheet_name

    wb.close()
    return mapping


# ============================================================
# GEOGRAPHY PARSING
# ============================================================

STATE_ABBREVS = ['NSW', 'VIC', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT']


def parse_geography(geo_value, geo_level='SA4'):
    """Parse geography string -> (geo_type, state, region_name) or (None,None,None).
    geo_level: 'SA4' or 'SA3' — determines how sub-state regions are typed.
    """
    if not geo_value or str(geo_value).strip() in ('', 'nan'):
        return None, None, None

    g = str(geo_value).strip()

    if g.upper() in ('AUSTRALIA', 'TOTAL', 'NATIONAL'):
        return 'Australia', None, 'Australia'

    if g in STATE_ABBREVS:
        return 'State', g, g

    if ' - ' in g:
        parts = g.split(' - ', 1)
        state = parts[0].strip()
        region = parts[1].strip()
        if state in STATE_ABBREVS:
            return geo_level, state, region

    return 'Unknown', None, g


# ============================================================
# TABLE PARSERS
# ============================================================

def find_header_row(df, keywords=None):
    """Find the row containing column headers."""
    if keywords is None:
        keywords = ['SA4 Region', 'State/Territory', 'SA3', 'Region']
    for idx in range(min(8, len(df))):
        row = df.iloc[idx]
        row_str = ' '.join([str(x) for x in row if not pd.isna(x)])
        if any(kw in row_str for kw in keywords):
            non_empty = sum(1 for x in row if not pd.isna(x) and str(x).strip())
            if non_empty >= 3:
                return idx
    return None


def parse_standard_table(filepath, sheet_name, geo_level='SA4'):
    """
    Parse a standard Supplement P table.
    geo_level: 'SA4' or 'SA3' — passed to parse_geography.
    """
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    except Exception as e:
        return None, "Could not read sheet: {}".format(e)

    header_row = find_header_row(df)
    if header_row is None:
        return None, "Could not find header row"

    # Extract headers, stop at first empty
    raw_headers = df.iloc[header_row]
    headers = []
    for h in raw_headers:
        if pd.isna(h) or str(h).strip() == '':
            break
        headers.append(str(h).strip())

    if len(headers) < 2:
        return None, "Too few columns ({})".format(len(headers))

    data_rows = df.iloc[header_row + 1:, :len(headers)].copy()
    data_rows.columns = headers

    first_col = headers[0]
    data_rows = data_rows.rename(columns={first_col: '_geo'})

    results = []
    for _, row in data_rows.iterrows():
        geo_type, state, region_name = parse_geography(row.get('_geo'), geo_level)
        if geo_type is None or geo_type == 'Unknown':
            continue

        for col in headers[1:]:
            if not col or col == '_geo':
                continue
            val = row.get(col)
            if pd.isna(val):
                continue
            try:
                val_float = float(val)
            except (ValueError, TypeError):
                continue
            results.append((geo_type, state, region_name, col, val_float))

    return results, None


def parse_p3_table(filepath, sheet_name):
    """Parse Table P.3 (Providers) which has a different header format."""
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    except Exception as e:
        return None, "Could not read: {}".format(e)

    header_row = None
    for idx in range(1, min(8, len(df))):
        cell0 = str(df.iloc[idx, 0]) if not pd.isna(df.iloc[idx, 0]) else ''
        if len(cell0) < 40 and ('State' in cell0 or 'Territory' in cell0):
            header_row = idx
            break

    if header_row is None:
        return None, "No P.3 header row found"

    num_cols = df.shape[1]
    col_headers = []
    for ci in range(1, num_cols):
        h = df.iloc[header_row, ci]
        hs = str(h).strip() if not pd.isna(h) else ''
        if not hs:
            break
        col_headers.append(hs)

    states = {'NSW', 'VIC', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT'}
    nationals = {'National', 'NATIONAL', 'Australia', 'AUSTRALIA', 'Total', 'TOTAL'}

    results = []
    for idx in range(header_row + 1, len(df)):
        cell0 = df.iloc[idx, 0]
        if pd.isna(cell0):
            continue
        geo_str = str(cell0).strip()
        if not geo_str:
            continue

        if geo_str in states:
            geo_type, state_abbr = 'State', geo_str
        elif geo_str in nationals or geo_str.lower() == 'national':
            geo_type, state_abbr = 'Australia', None
        else:
            continue

        region_name = 'Australia' if geo_type == 'Australia' else geo_str

        for ci, col_name in enumerate(col_headers):
            cell_val = df.iloc[idx, ci + 1] if ci + 1 < df.shape[1] else None
            if cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val)):
                continue
            s = str(cell_val).strip()
            if s.startswith('<'):
                val = 2.0  # midpoint for suppressed values
            else:
                try:
                    val = float(s.replace(',', ''))
                except ValueError:
                    continue
            results.append((geo_type, state_abbr, region_name, col_name, val))

    return results, None


def parse_p1_table(filepath, sheet_name):
    """Parse Table P.1 (Participants with SDA/SIL) — pivot-style table."""
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    except Exception as e:
        return None, "Could not read: {}".format(e)

    # P.1 typically has a different structure - cohort rows with columns for counts
    # For now, try the standard parser with relaxed keywords
    header_row = find_header_row(df, keywords=['Cohort', 'State', 'Active', 'SDA', 'SIL'])
    if header_row is None:
        return None, "Could not find P.1 header row"

    raw_headers = df.iloc[header_row]
    headers = []
    for h in raw_headers:
        if pd.isna(h) or str(h).strip() == '':
            break
        headers.append(str(h).strip())

    if len(headers) < 2:
        return None, "Too few columns"

    data_rows = df.iloc[header_row + 1:, :len(headers)].copy()
    data_rows.columns = headers

    first_col = headers[0]
    results = []
    for _, row in data_rows.iterrows():
        label = row[first_col]
        if pd.isna(label) or not str(label).strip():
            continue
        label_str = str(label).strip()
        for col in headers[1:]:
            val = row.get(col)
            if pd.isna(val):
                continue
            try:
                val_float = float(val)
            except (ValueError, TypeError):
                continue
            results.append(('Australia', None, 'Australia', '{} - {}'.format(label_str, col), val_float))

    return results, None



# ============================================================
# QUARTER HELPERS
# ============================================================

def quarter_ge(quarter_str, threshold):
    """Return True if quarter_str >= threshold (e.g. '2024-Q1' >= '2024-Q1')."""
    return quarter_str >= threshold


# ============================================================
# PARTICIPANT NEED PARSERS
# ============================================================

# Canonical measure names we store regardless of the raw column label
_TOTAL_MEASURE     = 'Total participants with SDA need'
_IN_USE_MEASURE    = 'Participants with SDA in use'
_ELIGIBLE_MEASURE  = 'Participants SDA eligible, not yet using SDA'

# The raw "total" column header changes across eras — identify it by position
# (always the last numeric column) or by keywords
_TOTAL_KEYWORDS = ['Total Participants', 'Total participants']

# Columns that carry the split cohort data (only clean from 2024-25 Q1)
_IN_USE_KEYWORDS      = ['with SDA in use']
_ELIGIBLE_KEYWORDS    = ['eligible, not yet using', 'SDA eligible']

# Design category columns for P.10 / P.18
_DC_COLUMNS = {
    'Improved Liveability': 'IL',
    'High Physical Support': 'HPS',
    'Robust': 'Robust',
    'Fully Accessible': 'FA',
    'Missing': 'Missing',
    'Basic': 'Basic',
}
_DC_TOTAL_KEYWORDS = ['Total participants with an SDA need', 'Total participants with SDA need']


def _match_col(headers, keywords):
    """Return first header that contains any keyword, or None."""
    for h in headers:
        if any(kw in h for kw in keywords):
            return h
    return None


def parse_participant_need_table(filepath, sheet_name, quarter_str, geo_level='SA4'):
    """
    Parse P.9 (SA4) or P.17 (SA3) participant need tables across all eras.

    Always loads: canonical total measure (all quarters)
    Also loads:   SDA in use + SDA eligible (2024-25 Q1 onward)

    Era rules:
      2022-23 Q2       total col = 'Total Participants with SDA need'
      2022-23 Q3 to 2023-24 Q4  total col = 'Total Participants with SDA funding or an SDA need'
      2024-25 Q1+      total col = 'Total Participants with SDA in use or SDA eligible...'
                       + split cols available
    """
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    except Exception as e:
        return None, "Could not read: {}".format(e)

    header_row = find_header_row(df, keywords=['SA4 Region', 'SA3 Region', 'Region'])
    if header_row is None:
        return None, "Could not find header row"

    raw_headers = df.iloc[header_row]
    headers = []
    for h in raw_headers:
        s = str(h).strip() if not pd.isna(h) else ''
        if not s:
            break
        headers.append(s)

    if len(headers) < 2:
        return None, "Too few columns"

    # Identify columns
    total_col    = _match_col(headers[1:], _TOTAL_KEYWORDS)
    in_use_col   = _match_col(headers[1:], _IN_USE_KEYWORDS)
    eligible_col = _match_col(headers[1:], _ELIGIBLE_KEYWORDS)

    if not total_col:
        # Fallback: last column
        total_col = headers[-1]

    load_split = quarter_ge(quarter_str, '2024-Q1') and in_use_col and eligible_col

    data_rows = df.iloc[header_row + 1:].copy()

    results = []
    for _, row in data_rows.iterrows():
        geo_raw = row.iloc[0]
        geo_type, state, region_name = parse_geography(geo_raw, geo_level)
        if geo_type is None or geo_type == 'Unknown':
            continue

        # Total — always
        t_idx = headers.index(total_col)
        t_val = row.iloc[t_idx]
        try:
            results.append((geo_type, state, region_name, _TOTAL_MEASURE, float(t_val)))
        except (ValueError, TypeError):
            pass

        # Split cohorts — 2024-25 Q1+
        if load_split:
            for col, measure in [(in_use_col, _IN_USE_MEASURE), (eligible_col, _ELIGIBLE_MEASURE)]:
                try:
                    idx = headers.index(col)
                    results.append((geo_type, state, region_name, measure, float(row.iloc[idx])))
                except (ValueError, TypeError, ValueError):
                    pass

    return results, None


def parse_participant_dc_table(filepath, sheet_name, quarter_str, geo_level='SA4'):
    """
    Parse P.10 (SA4) or P.18 (SA3) participant need by Design Category.
    Only called for 2024-25 Q2 onward — earlier quarters are skipped upstream.
    Stores each DC as a canonical measure name.
    """
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    except Exception as e:
        return None, "Could not read: {}".format(e)

    header_row = find_header_row(df, keywords=['SA4 Region', 'SA3 Region', 'Region'])
    if header_row is None:
        return None, "Could not find header row"

    raw_headers = df.iloc[header_row]
    headers = []
    for h in raw_headers:
        s = str(h).strip() if not pd.isna(h) else ''
        if not s:
            break
        headers.append(s)

    if len(headers) < 2:
        return None, "Too few columns"

    # Map raw header -> canonical measure name
    col_map = {}
    for h in headers[1:]:
        for raw_dc, canonical in _DC_COLUMNS.items():
            if raw_dc in h:
                col_map[h] = 'DC: ' + canonical
                break
        if h not in col_map:
            # Check total
            if any(kw in h for kw in _DC_TOTAL_KEYWORDS):
                col_map[h] = 'DC: Total'

    if not col_map:
        return None, "No recognisable DC columns found"

    data_rows = df.iloc[header_row + 1:].copy()
    results = []

    for _, row in data_rows.iterrows():
        geo_raw = row.iloc[0]
        geo_type, state, region_name = parse_geography(geo_raw, geo_level)
        if geo_type is None or geo_type == 'Unknown':
            continue

        for col_idx, col in enumerate(headers[1:], start=1):
            measure = col_map.get(col)
            if not measure:
                continue
            try:
                val = float(row.iloc[col_idx])
                results.append((geo_type, state, region_name, measure, val))
            except (ValueError, TypeError):
                pass

    return results, None


def parse_p2_financial_table(filepath, sheet_name):
    """
    Parse Table P.2 — Annualised committed supports (SDA and SIL).
    State-level only. Strips $ and commas from early-quarter string values.
    Skips percentage columns (dashboard calculates these).
    Stores 3 measures: SDA committed, SIL committed, Total committed.
    """
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    except Exception as e:
        return None, "Could not read: {}".format(e)

    header_row = find_header_row(df, keywords=['State/Territory', 'State', 'Territory'])
    if header_row is None:
        return None, "Could not find header row"

    raw_headers = df.iloc[header_row]
    headers = []
    for h in raw_headers:
        s = str(h).strip() if not pd.isna(h) else ''
        if not s:
            break
        headers.append(s)

    # Identify the 3 dollar-value columns by keyword, skip % columns
    SDA_KEYWORDS   = ['committed to SDA', 'SDA in current']
    SIL_KEYWORDS   = ['committed support', 'SIL']
    TOTAL_KEYWORDS = ['Total committed', 'Total committed in support']

    sda_col   = _match_col(headers[1:], SDA_KEYWORDS)
    sil_col   = _match_col(headers[1:], SIL_KEYWORDS)
    total_col = _match_col(headers[1:], TOTAL_KEYWORDS)

    def clean_dollar(v):
        """Strip $, commas; return float or raise."""
        s = str(v).strip().replace('$', '').replace(',', '')
        return float(s)

    STATES = {'NSW', 'VIC', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT'}
    TOTALS = {'Total', 'TOTAL', 'Australia', 'AUSTRALIA'}

    data_rows = df.iloc[header_row + 1:].copy()
    results = []

    for _, row in data_rows.iterrows():
        geo_raw = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
        if not geo_raw:
            continue

        if geo_raw in STATES:
            geo_type, state, region_name = 'State', geo_raw, geo_raw
        elif geo_raw in TOTALS:
            geo_type, state, region_name = 'Australia', None, 'Australia'
        else:
            continue  # skip Other Territories, Missing, footnotes

        for col, measure in [
            (sda_col,   'SDA Committed ($)'),
            (sil_col,   'SIL Committed ($)'),
            (total_col, 'Total Committed ($)'),
        ]:
            if not col:
                continue
            try:
                idx = headers.index(col)
                val = clean_dollar(row.iloc[idx])
                results.append((geo_type, state, region_name, measure, val))
            except (ValueError, TypeError, IndexError):
                pass

    return results, None


# ============================================================
# TABLE P.6 — SDA CAPACITY EXTRACTION
# ============================================================

def extract_sda_capacity(conn, filepath, sheet_name, quarter_str):
    """
    Extract total SDA capacity from Table P.6 'Total' row.

    Column layout (consistent across all eras):
      Col 0: Geography label
      Col 1: 1 Resident dwellings
      Col 2: 2 Residents dwellings
      Col 3: 3 Residents dwellings
      Col 4: 4 Residents dwellings
      Col 5: 5 Residents dwellings
      Col 6: 6+ Residents dwellings   (treated as 6 for capacity calc)
      Col 7: Total dwellings

    total_places = 1*d1 + 2*d2 + 3*d3 + 4*d4 + 5*d5 + 6*d6

    Upserts one row per quarter into sda_capacity.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name]

        total_row = None
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and str(row[0]).strip() == 'Total':
                total_row = list(row)
                break
        wb.close()

        if total_row is None:
            print("    {} Table P.6: 'Total' row not found — capacity extraction skipped".format(
                chr(9888)))
            return

        def _safe_int(v):
            try:
                return int(float(str(v).replace(',', '').strip()))
            except (ValueError, TypeError):
                return 0

        d1 = _safe_int(total_row[1])  # 1 Resident
        d2 = _safe_int(total_row[2])  # 2 Residents
        d3 = _safe_int(total_row[3])  # 3 Residents
        d4 = _safe_int(total_row[4])  # 4 Residents
        d5 = _safe_int(total_row[5])  # 5 Residents
        d6 = _safe_int(total_row[6])  # 6+ Residents (use 6 as multiplier)
        total_dwellings = _safe_int(total_row[7])
        total_places = 1*d1 + 2*d2 + 3*d3 + 4*d4 + 5*d5 + 6*d6

        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO sda_capacity (quarter, total_dwellings, total_places)
            VALUES (?, ?, ?)
        """, (quarter_str, total_dwellings, total_places))
        conn.commit()
        print("    {} Table P.6 capacity: {:,} dwellings, {:,} places".format(
            chr(10003), total_dwellings, total_places))

    except Exception as e:
        print("    {} Table P.6 capacity extraction error: {}".format(chr(9888), e))


# ============================================================
# DATABASE
# ============================================================

def create_fresh_database(db_path):
    """Wipe and recreate database."""
    if os.path.exists(db_path):
        os.remove(db_path)
        print("Removed old database")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.executescript("""
        CREATE TABLE quarters (
            quarter_id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL,
            quarter_num INTEGER NOT NULL,
            quarter_str TEXT NOT NULL UNIQUE,
            UNIQUE(year, quarter_num)
        );
        CREATE TABLE datasets (
            dataset_id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_table TEXT NOT NULL UNIQUE,
            description TEXT
        );
        CREATE TABLE geographies (
            geo_id INTEGER PRIMARY KEY AUTOINCREMENT,
            geo_type TEXT NOT NULL,
            name TEXT NOT NULL,
            state TEXT,
            parent_id INTEGER REFERENCES geographies(geo_id),
            UNIQUE(geo_type, name, state)
        );
        CREATE TABLE measures (
            measure_id INTEGER PRIMARY KEY AUTOINCREMENT,
            quarter_id INTEGER NOT NULL REFERENCES quarters(quarter_id),
            dataset_id INTEGER NOT NULL REFERENCES datasets(dataset_id),
            geo_id INTEGER NOT NULL REFERENCES geographies(geo_id),
            measure_name TEXT,
            value REAL,
            UNIQUE(quarter_id, dataset_id, geo_id, measure_name)
        );
        CREATE INDEX idx_measures_quarter ON measures(quarter_id);
        CREATE INDEX idx_measures_dataset ON measures(dataset_id);
        CREATE INDEX idx_measures_geo ON measures(geo_id);
        CREATE INDEX idx_measures_composite ON measures(quarter_id, dataset_id, geo_id);

        CREATE TABLE IF NOT EXISTS sda_capacity (
            quarter TEXT PRIMARY KEY,
            total_dwellings INTEGER,
            total_places INTEGER
        );

        INSERT INTO geographies (geo_type, name, state, parent_id)
        VALUES ('Australia', 'Australia', NULL, NULL);
        INSERT INTO geographies (geo_type, name, state, parent_id) VALUES
        ('State', 'New South Wales', 'NSW', 1),
        ('State', 'Victoria', 'VIC', 1),
        ('State', 'Queensland', 'QLD', 1),
        ('State', 'South Australia', 'SA', 1),
        ('State', 'Western Australia', 'WA', 1),
        ('State', 'Tasmania', 'TAS', 1),
        ('State', 'Northern Territory', 'NT', 1),
        ('State', 'Australian Capital Territory', 'ACT', 1);
    """)
    conn.commit()
    conn.close()
    print("Fresh database created: {}".format(db_path))


def get_or_create_quarter(conn, quarter_str):
    cursor = conn.cursor()
    cursor.execute("SELECT quarter_id FROM quarters WHERE quarter_str=?", (quarter_str,))
    row = cursor.fetchone()
    if row:
        return row[0]
    year, q = quarter_str.split('-Q')
    cursor.execute("INSERT INTO quarters (year, quarter_num, quarter_str) VALUES (?,?,?)",
                   (int(year), int(q), quarter_str))
    conn.commit()
    return cursor.lastrowid


def get_or_create_geography(conn, geo_type, name, state):
    cursor = conn.cursor()
    if state:
        cursor.execute("SELECT geo_id FROM geographies WHERE geo_type=? AND name=? AND state=?",
                       (geo_type, name, state))
    else:
        cursor.execute("SELECT geo_id FROM geographies WHERE geo_type=? AND name=? AND state IS NULL",
                       (geo_type, name))
    row = cursor.fetchone()
    if row:
        return row[0]

    parent_id = None
    if geo_type in ('SA4', 'SA3') and state:
        cursor.execute("SELECT geo_id FROM geographies WHERE geo_type='State' AND state=?", (state,))
        row = cursor.fetchone()
        parent_id = row[0] if row else None

    cursor.execute("INSERT OR IGNORE INTO geographies (geo_type, name, state, parent_id) VALUES (?,?,?,?)",
                   (geo_type, name, state, parent_id))
    conn.commit()
    cursor.execute("SELECT geo_id FROM geographies WHERE geo_type=? AND name=? AND state IS ?",
                   (geo_type, name, state))
    return cursor.fetchone()[0]


def get_dataset_id(conn, source_table, description=''):
    cursor = conn.cursor()
    cursor.execute("SELECT dataset_id FROM datasets WHERE source_table=?", (source_table,))
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute("INSERT OR IGNORE INTO datasets (source_table, description) VALUES (?,?)",
                   (source_table, description))
    conn.commit()
    cursor.execute("SELECT dataset_id FROM datasets WHERE source_table=?", (source_table,))
    return cursor.fetchone()[0]


# ============================================================
# FILE DISCOVERY
# ============================================================

def parse_quarter_from_filename(filename):
    """Extract quarter from filename. Handles multiple naming conventions."""
    # '2025-26 Q2' or '2025-26_Q2'
    match = re.search(r'(\d{4})-\d{2}[\s_]+Q(\d)', filename)
    if match:
        return '{}-Q{}'.format(match.group(1), match.group(2))
    # '202526_Q2'
    match = re.search(r'(\d{4})\d{2}_Q(\d)', filename)
    if match:
        return '{}-Q{}'.format(match.group(1), match.group(2))
    return None


def find_supplement_files(folder_path):
    """Find all Supplement P files and deduplicate by quarter."""
    folder = Path(folder_path)
    files = sorted([
        f for f in folder.glob("*.xlsx")
        if 'supplement' in f.name.lower()
        and re.search(r'\d{4}', f.name)
        and 'mapping' not in f.name.lower()
        and 'evolution' not in f.name.lower()
        and 'BACKUP' not in f.name
    ])

    # Deduplicate: prefer files without '(1)' or '_0' in name
    seen = {}
    for f in files:
        q = parse_quarter_from_filename(f.name)
        if not q:
            continue
        if q not in seen:
            seen[q] = f
        else:
            # Prefer cleaner filename
            existing = seen[q]
            if '(1)' in existing.name or '_0' in existing.name:
                seen[q] = f

    return sorted(seen.items())


# ============================================================
# MAIN LOADER
# ============================================================

def load_file(conn, filepath, quarter_str):
    """Load all detected tables from one Supplement P file."""
    print("\n  Loading {}: {}".format(quarter_str, Path(filepath).name))

    # Detect sheets by title
    try:
        sheet_map = detect_sheets(filepath)
    except Exception as e:
        print("    ERROR opening file: {}".format(e))
        return 0

    if not sheet_map:
        print("    No table sheets detected")
        return 0

    quarter_id = get_or_create_quarter(conn, quarter_str)
    total_inserted = 0
    cursor = conn.cursor()

    for canonical, sheet_name in sorted(sheet_map.items()):
        # Choose parser based on table type
        if canonical == 'Table P.3':
            records, error = parse_p3_table(str(filepath), sheet_name)

        elif canonical == 'Table P.1':
            records, error = parse_p1_table(str(filepath), sheet_name)

        elif canonical in ('Table P.9',):
            # Participant need SA4 — all quarters, era-aware
            records, error = parse_participant_need_table(
                str(filepath), sheet_name, quarter_str, geo_level='SA4')

        elif canonical in ('Table P.17',):
            # Participant need SA3 — all quarters, era-aware
            records, error = parse_participant_need_table(
                str(filepath), sheet_name, quarter_str, geo_level='SA3')

        elif canonical in ('Table P.10',):
            # Participant DC SA4 — only 2024-25 Q2 onward
            if not quarter_ge(quarter_str, '2024-Q2'):
                print("    -- {} -> {}: skipped (era: pre 2024-25 Q2)".format(
                    sheet_name, canonical))
                continue
            records, error = parse_participant_dc_table(
                str(filepath), sheet_name, quarter_str, geo_level='SA4')

        elif canonical in ('Table P.18',):
            # Participant DC SA3 — only 2024-25 Q2 onward
            if not quarter_ge(quarter_str, '2024-Q2'):
                print("    -- {} -> {}: skipped (era: pre 2024-25 Q2)".format(
                    sheet_name, canonical))
                continue
            records, error = parse_participant_dc_table(
                str(filepath), sheet_name, quarter_str, geo_level='SA3')

        elif canonical == 'Table P.2':
            # Financial — all quarters
            records, error = parse_p2_financial_table(str(filepath), sheet_name)

        elif canonical == 'Table P.6':
            # Enrolled dwellings by max residents — standard parse into measures table
            records, error = parse_standard_table(str(filepath), sheet_name)
            # Also extract capacity summary into sda_capacity table
            if not error:
                extract_sda_capacity(conn, str(filepath), sheet_name, quarter_str)

        else:
            records, error = parse_standard_table(str(filepath), sheet_name)

        if error:
            print("    {} {} -> {}: {}".format(
                chr(9888), sheet_name, canonical, error))
            continue

        if not records:
            print("    {} {} -> {}: No data".format(
                chr(9888), sheet_name, canonical))
            continue

        # Find description from signatures
        desc = ''
        for sig in SHEET_SIGNATURES:
            if sig['canonical'] == canonical:
                desc = sig['desc']
                break

        dataset_id = get_dataset_id(conn, canonical, desc)
        inserted = 0

        for geo_type, state, region_name, measure_name, value in records:
            geo_id = get_or_create_geography(conn, geo_type, region_name, state)
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO measures
                    (quarter_id, dataset_id, geo_id, measure_name, value)
                    VALUES (?,?,?,?,?)
                """, (quarter_id, dataset_id, geo_id, measure_name, value))
                inserted += 1
            except Exception:
                pass

        conn.commit()
        marker = ' ***' if canonical in ('Table P.11', 'Table P.12') else ''
        print("    {} {} -> {}: {:,} measures{}".format(
            chr(10003), sheet_name, canonical, inserted, marker))
        total_inserted += inserted

    return total_inserted


def verify_database(db_path):
    """Print verification summary."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*), MIN(quarter_str), MAX(quarter_str) FROM quarters")
    q_count, min_q, max_q = cursor.fetchone()
    print("\nQuarters: {} ({} to {})".format(q_count, min_q, max_q))

    cursor.execute("SELECT COUNT(DISTINCT geo_id) FROM measures")
    print("Geographies: {}".format(cursor.fetchone()[0]))

    cursor.execute("SELECT COUNT(*) FROM measures")
    print("Total Measures: {:,}".format(cursor.fetchone()[0]))

    # Measures by dataset
    print("\nMeasures by dataset:")
    cursor.execute("""
        SELECT d.source_table, d.description, COUNT(*) as cnt
        FROM measures m
        JOIN datasets d ON m.dataset_id = d.dataset_id
        GROUP BY d.source_table
        ORDER BY d.source_table
    """)
    for row in cursor.fetchall():
        print("  {:12s} {:,>8d}  {}".format(row[0], row[2], row[1][:50]))

    # Key validation: total enrolled dwellings
    print("\nTotal Enrolled Dwellings over time (P.4 Total, Australia):")
    cursor.execute("""
        SELECT q.quarter_str, m.value
        FROM measures m
        JOIN quarters q ON m.quarter_id = q.quarter_id
        JOIN datasets d ON m.dataset_id = d.dataset_id
        JOIN geographies g ON m.geo_id = g.geo_id
        WHERE d.source_table = 'Table P.4'
          AND g.geo_type = 'Australia'
          AND m.measure_name IN ('TOTAL', 'Total')
        ORDER BY q.year, q.quarter_num
    """)
    for row in cursor.fetchall():
        print("  {}: {:,.0f}".format(row[0], row[1]))

    # Participant need totals (P.9, Australia)
    print("\nTotal Participants with SDA Need over time (P.9, Australia):")
    cursor.execute("""
        SELECT q.quarter_str, m.value
        FROM measures m
        JOIN quarters q ON m.quarter_id = q.quarter_id
        JOIN datasets d ON m.dataset_id = d.dataset_id
        JOIN geographies g ON m.geo_id = g.geo_id
        WHERE d.source_table = 'Table P.9'
          AND g.geo_type = 'Australia'
          AND m.measure_name = 'Total participants with SDA need'
        ORDER BY q.year, q.quarter_num
    """)
    for row in cursor.fetchall():
        print("  {}: {:,.0f}".format(row[0], row[1]))

    # Missing cohort trend (P.10, Australia)
    print("\n'Missing' DC cohort trend over time (P.10, Australia):")
    cursor.execute("""
        SELECT q.quarter_str, m.value
        FROM measures m
        JOIN quarters q ON m.quarter_id = q.quarter_id
        JOIN datasets d ON m.dataset_id = d.dataset_id
        JOIN geographies g ON m.geo_id = g.geo_id
        WHERE d.source_table = 'Table P.10'
          AND g.geo_type = 'Australia'
          AND m.measure_name = 'DC: Missing'
        ORDER BY q.year, q.quarter_num
    """)
    for row in cursor.fetchall():
        print("  {}: {:,.0f}".format(row[0], row[1]))

    # Financial trend (P.2, Australia)
    print("\nSDA Committed Supports over time (P.2, Australia $):")
    cursor.execute("""
        SELECT q.quarter_str, m.value
        FROM measures m
        JOIN quarters q ON m.quarter_id = q.quarter_id
        JOIN datasets d ON m.dataset_id = d.dataset_id
        JOIN geographies g ON m.geo_id = g.geo_id
        WHERE d.source_table = 'Table P.2'
          AND g.geo_type = 'Australia'
          AND m.measure_name = 'SDA Committed ($)'
        ORDER BY q.year, q.quarter_num
    """)
    for row in cursor.fetchall():
        print("  {}: ${:,.0f}".format(row[0], row[1]))

    # P.11 measures per quarter (should be ~7000 each)
    print("\nP.11 (New Build multi-dim) measures per quarter:")
    cursor.execute("""
        SELECT q.quarter_str, COUNT(*) as cnt
        FROM measures m
        JOIN quarters q ON m.quarter_id = q.quarter_id
        JOIN datasets d ON m.dataset_id = d.dataset_id
        WHERE d.source_table = 'Table P.11'
        GROUP BY q.quarter_str
        ORDER BY q.year, q.quarter_num
    """)
    for row in cursor.fetchall():
        flag = ' !!!' if row[1] < 5000 else ''
        print("  {}: {:,}{}".format(row[0], row[1], flag))

    # P.12 measures per quarter
    print("\nP.12 (Existing multi-dim) measures per quarter:")
    cursor.execute("""
        SELECT q.quarter_str, COUNT(*) as cnt
        FROM measures m
        JOIN quarters q ON m.quarter_id = q.quarter_id
        JOIN datasets d ON m.dataset_id = d.dataset_id
        WHERE d.source_table = 'Table P.12'
        GROUP BY q.quarter_str
        ORDER BY q.year, q.quarter_num
    """)
    for row in cursor.fetchall():
        flag = ' !!!' if row[1] < 4000 else ''
        print("  {}: {:,}{}".format(row[0], row[1], flag))

    conn.close()


def load_all(db_path, folder_path):
    """Main entry point."""
    print("=" * 60)
    print("SDA Dashboard - v4 Title-Based Loader")
    print("=" * 60)

    # Find files
    file_quarters = find_supplement_files(folder_path)
    if not file_quarters:
        print("\nNo Supplement P files found in {}".format(folder_path))
        return

    print("\nFound {} quarters to load:".format(len(file_quarters)))
    for q, f in file_quarters:
        print("  {}: {}".format(q, f.name))

    # Create fresh database
    print("\n" + "=" * 60)
    create_fresh_database(db_path)

    # Load each file
    conn = sqlite3.connect(db_path)
    grand_total = 0
    results = []

    for quarter_str, filepath in file_quarters:
        count = load_file(conn, str(filepath), quarter_str)
        grand_total += count
        results.append((quarter_str, count))

    conn.close()

    # Summary
    print("\n" + "=" * 60)
    print("LOAD COMPLETE")
    print("=" * 60)
    for q, count in results:
        status = chr(10003) if count > 0 else chr(10007)
        print("  {} {}: {:,} measures".format(status, q, count))

    print("\nGrand Total: {:,} measures".format(grand_total))

    # Verify
    print("\n" + "=" * 60)
    print("VERIFICATION")
    print("=" * 60)
    verify_database(db_path)


# ============================================================
# IMPORTABLE ENTRY POINT  (called by app.py admin upload)
# ============================================================

def run_loader(filepath, quarter_str, mode, db_path=None):
    """
    Load a single Supplement P Excel file into the database.
    Designed to be imported and called from app.py for the admin upload feature.

    Args:
        filepath (str):      Path to the .xlsx file to load.
        quarter_str (str):   Quarter in calendar 'YYYY-QN' format (e.g. '2025-Q3').
        mode (str):          'new'     — insert rows, skip if quarter already exists.
                             'replace' — delete all existing rows for this quarter first,
                                         then reload from the file.
        db_path (str|None):  Path to the SQLite DB.
                             Resolved in order:
                               1. db_path argument (if provided)
                               2. DB_PATH environment variable
                               3. sda_dashboard_clean.db next to this module

    Returns:
        dict: {
            'success':     bool,
            'quarter':     str,
            'rows_loaded': int,
            'message':     str,
        }
    """
    # 1 — Resolve DB path
    if db_path is None:
        db_path = os.environ.get(
            'DB_PATH',
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sda_dashboard_clean.db')
        )

    # 2 — Validate quarter_str format
    if not re.match(r'^\d{4}-Q[1-4]$', quarter_str):
        return {
            'success': False, 'quarter': quarter_str, 'rows_loaded': 0,
            'message': 'Invalid quarter format: {!r}. Expected YYYY-QN (e.g. 2025-Q3).'.format(quarter_str),
        }

    # 3 — DB must already exist (do not wipe it)
    if not os.path.exists(db_path):
        return {
            'success': False, 'quarter': quarter_str, 'rows_loaded': 0,
            'message': 'Database not found: {}'.format(db_path),
        }

    conn = None
    try:
        conn = sqlite3.connect(db_path)

        # Ensure sda_capacity table exists (for existing DBs predating this feature)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sda_capacity (
                quarter TEXT PRIMARY KEY,
                total_dwellings INTEGER,
                total_places INTEGER
            )
        """)
        conn.commit()

        # 4 — Replace mode: delete existing measures and capacity for this quarter
        if mode == 'replace':
            cursor = conn.cursor()
            cursor.execute("SELECT quarter_id FROM quarters WHERE quarter_str=?", (quarter_str,))
            row = cursor.fetchone()
            if row:
                cursor.execute("DELETE FROM measures WHERE quarter_id=?", (row[0],))
                cursor.execute("DELETE FROM sda_capacity WHERE quarter=?", (quarter_str,))
                conn.commit()
                print("  Deleted existing measures for {} (replace mode)".format(quarter_str))

        # 5 — Load the file
        rows = load_file(conn, str(filepath), quarter_str)
        conn.close()
        conn = None

        return {
            'success': True,
            'quarter': quarter_str,
            'rows_loaded': rows,
            'message': 'Successfully loaded {:,} measures for {}.'.format(rows, quarter_str),
        }

    except Exception as exc:
        if conn:
            try:
                conn.close()
            except Exception:
                pass
        return {
            'success': False,
            'quarter': quarter_str,
            'rows_loaded': 0,
            'message': str(exc),
        }


if __name__ == '__main__':
    _default_db = os.environ.get(
        'DB_PATH',
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sda_dashboard_clean.db')
    )
    parser = argparse.ArgumentParser(description='SDA Dashboard Data Loader v4')
    parser.add_argument('--folder', default=r'C:\Users\grend\Downloads',
                        help='Folder containing Supplement P Excel files')
    parser.add_argument('--db', default=_default_db,
                        help='Database path (overrides DB_PATH env var)')
    parser.add_argument('--verify', action='store_true',
                        help='Just verify existing database')
    args = parser.parse_args()

    if args.verify:
        verify_database(args.db)
    else:
        load_all(args.db, args.folder)
